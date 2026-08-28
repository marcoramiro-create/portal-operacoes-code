from __future__ import annotations

import json
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def trim_text(value: str) -> str:
    return " ".join(value.split())


def analyze(path: Path) -> dict:
    workbook = load_workbook(path, read_only=False, data_only=False)
    sheets = []
    for worksheet in workbook.worksheets:
        merged_ranges = [str(item) for item in worksheet.merged_cells.ranges]
        hidden_rows = [index for index, dimension in worksheet.row_dimensions.items() if dimension.hidden]
        hidden_columns = [key for key, dimension in worksheet.column_dimensions.items() if dimension.hidden]
        formulas = 0
        trimmed_cells = 0
        internal_space_cells = 0
        value_types = Counter()
        nonempty_by_row: list[tuple[int, int]] = []

        for row in worksheet.iter_rows():
            nonempty = 0
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                nonempty += 1
                value_types[type(value).__name__] += 1
                if isinstance(value, str):
                    if value != value.strip():
                        trimmed_cells += 1
                    if trim_text(value) != value.strip():
                        internal_space_cells += 1
                    if value.startswith("="):
                        formulas += 1
            nonempty_by_row.append((row[0].row, nonempty))

        header_candidates = sorted(nonempty_by_row, key=lambda pair: pair[1], reverse=True)[:5]
        header_rows = []
        for row_number, count in header_candidates:
            values = [worksheet.cell(row_number, column).value for column in range(1, worksheet.max_column + 1)]
            header_rows.append({"row": row_number, "nonempty": count, "values": values})

        data_header_row = next(
            (
                row_number
                for row_number, _ in nonempty_by_row
                if str(worksheet.cell(row_number, 1).value or "").strip().upper() == "FILIAL"
            ),
            None,
        )
        data_rows = []
        filial_counts: Counter[str] = Counter()
        code_pairs: Counter[tuple[str, str, str]] = Counter()
        date_columns: list[dict] = []

        if data_header_row:
            for column in range(1, worksheet.max_column + 1):
                raw = worksheet.cell(data_header_row, column).value
                normalized = trim_text(str(raw)) if raw is not None else ""
                if normalized.isdigit() and len(normalized) == 8:
                    try:
                        parsed = datetime.strptime(normalized, "%Y%m%d")
                    except ValueError:
                        pass
                    else:
                        date_columns.append({"column": column, "raw": normalized, "iso": parsed.date().isoformat()})

            for row_number in range(data_header_row + 1, worksheet.max_row + 1):
                filial = trim_text(str(worksheet.cell(row_number, 1).value or ""))
                aggregate_code = trim_text(str(worksheet.cell(row_number, 2).value or ""))
                code = trim_text(str(worksheet.cell(row_number, 3).value or ""))
                if not filial and not aggregate_code and not code:
                    continue
                data_rows.append(row_number)
                filial_counts[filial] += 1
                code_pairs[(filial, aggregate_code, code)] += 1

        duplicate_keys = [
            {"filial": key[0], "aggregate_code": key[1], "code": key[2], "count": count}
            for key, count in code_pairs.items()
            if count > 1
        ]
        sheets.append(
            {
                "name": worksheet.title,
                "dimensions": {"rows": worksheet.max_row, "columns": worksheet.max_column},
                "merged_ranges": merged_ranges,
                "hidden_rows": hidden_rows,
                "hidden_columns": hidden_columns,
                "formula_cells": formulas,
                "text_cells_with_edge_spaces": trimmed_cells,
                "text_cells_with_repeated_internal_spaces": internal_space_cells,
                "value_types": dict(value_types),
                "header_candidates": header_rows,
                "data_header_row": data_header_row,
                "data_row_count": len(data_rows),
                "filial_counts": dict(filial_counts),
                "date_columns": date_columns,
                "duplicate_business_keys": duplicate_keys[:20],
                "duplicate_business_key_count": len(duplicate_keys),
            }
        )

    return {"file": path.name, "sheets": sheets}


def main() -> None:
    paths = [Path(item) for item in sys.argv[1:]]
    print(json.dumps([analyze(path) for path in paths], ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
